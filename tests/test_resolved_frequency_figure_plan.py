from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from sciplot_core.figure_plan import (
    FigurePlanResolutionError,
    resolve_current_figure_plan,
    resolve_figure_plan,
    resolve_preparation_figure_plan,
)
from sciplot_core.studio_core.figure_requests import (
    _rheology_frequency_figure_queue,
)
from sciplot_core.studio_core.publish_sources import (
    StudioRunSources,
    verify_studio_run_source_binding,
)
from sciplot_core.study_model import (
    STUDY_MODEL_KIND,
    STUDY_MODEL_VERSION,
    experiment_recommendation_payload,
)
from sciplot_core.workflow.rheology_bundle import _sweep_metric_sources


def _frequency_study_model() -> dict[str, object]:
    recommendation = experiment_recommendation_payload(
        rule_id="rheology_frequency_sweep"
    )
    return {
        "kind": STUDY_MODEL_KIND,
        "version": STUDY_MODEL_VERSION,
        "figure_queue": [
            {**figure, "order": order, "status": "planned"}
            for order, figure in enumerate(
                recommendation["figure_queue"],
                start=1,
            )
        ],
    }


def _write_frequency_source(
    path: Path,
    *,
    include_complex_modulus: bool,
    include_loss_modulus: bool = False,
    include_loss_factor_and_viscosity: bool = False,
) -> None:
    headers = ["Angular Frequency", "Storage Modulus"]
    samples = ["Sample A", "Sample A"]
    units = ["rad/s", "Pa"]
    values: list[list[object]] = [[1.0, 100.0], [10.0, 120.0]]
    if include_loss_modulus:
        headers.append("Loss Modulus")
        samples.append("Sample A")
        units.append("Pa")
        for row, value in zip(values, (80.0, 90.0), strict=True):
            row.append(value)
    if include_loss_factor_and_viscosity:
        headers.extend(["Loss Factor", "Complex Viscosity"])
        samples.extend(["Sample A", "Sample A"])
        units.extend(["1", "mPa·s"])
        for row, loss_factor, viscosity in zip(
            values,
            (0.8, 0.75),
            (128_000.0, 15_000.0),
            strict=True,
        ):
            row.extend([loss_factor, viscosity])
    if include_complex_modulus:
        headers.append("Complex Modulus")
        samples.append("Sample A")
        units.append("Pa")
        for row, value in zip(values, (140.0, 160.0), strict=True):
            row.append(value)
    pd.DataFrame([headers, samples, units, *values]).to_excel(
        path,
        header=False,
        index=False,
    )


def _write_frequency_text_source(path: Path) -> None:
    path.write_text(
        "Angular Frequency,Storage Modulus\n"
        "rad/s,Pa\n"
        "1,100\n"
        "10,90\n",
        encoding="utf-8",
    )


def test_frequency_plan_keeps_source_available_complex_modulus(
    tmp_path: Path,
) -> None:
    source = tmp_path / "frequency.xlsx"
    _write_frequency_source(source, include_complex_modulus=True)
    study_model = _frequency_study_model()

    plan = resolve_figure_plan(
        rule_id="rheology_frequency_sweep",
        template="point_line",
        study_model=study_model,
        input_path=source,
        request={},
    )

    assert plan is not None
    assert plan.selected_figure_ids == (
        "storage_modulus_vs_frequency",
        "loss_modulus_vs_frequency",
        "loss_factor_vs_frequency",
        "complex_viscosity_vs_frequency",
        "complex_modulus_vs_frequency",
    )
    extra = plan.tasks[-1]
    assert extra.y_metric == "complex_modulus"
    assert extra.artifact_stem == "freq_complex_modulus"
    assert extra.document_stem == "complex_modulus_vs_frequency"

    request = {
        "rule_id": "rheology_frequency_sweep",
        "template": "point_line",
        "study_model": study_model,
        "resolved_figure_plan": plan.to_payload(),
    }
    studio_queue = _rheology_frequency_figure_queue(
        request,
        figure_plan=plan,
    )
    workflow_sources = _sweep_metric_sources(
        source,
        request=request,
        output_dir=tmp_path / "run",
    )

    assert [item["id"] for item in studio_queue] == list(plan.selected_figure_ids)
    assert [item[0] for item in workflow_sources] == [
        "freq_storage_modulus",
        "freq_complex_modulus",
    ]


def test_frequency_plan_without_extra_metric_keeps_default_four(
    tmp_path: Path,
) -> None:
    source = tmp_path / "frequency.xlsx"
    _write_frequency_source(source, include_complex_modulus=False)

    plan = resolve_figure_plan(
        rule_id="rheology_frequency_sweep",
        template="point_line",
        study_model=_frequency_study_model(),
        input_path=source,
        request={},
    )

    assert plan is not None
    assert len(plan.tasks) == 4
    assert "complex_modulus_vs_frequency" not in plan.selected_figure_ids


def test_frequency_directory_plan_uses_parser_selected_text_sources(
    tmp_path: Path,
) -> None:
    source = tmp_path / "frequency"
    source.mkdir()
    raw = source / "sample_a.csv"
    _write_frequency_text_source(raw)
    (source / "derived.xlsx").write_bytes(b"must not become planning evidence")

    plan = resolve_figure_plan(
        rule_id="rheology_frequency_sweep",
        template="point_line",
        study_model=_frequency_study_model(),
        input_path=source,
        request={},
    )

    assert plan is not None
    assert plan.selected_figure_ids == ("storage_modulus_vs_frequency",)
    assert all(task.sample_order == ("sample_a",) for task in plan.tasks)
    assert all(task.replicate_counts == (("sample_a", 1),) for task in plan.tasks)


def test_frequency_queue_fallback_does_not_scan_process_working_directory(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    unrelated = tmp_path / "unrelated.xlsx"
    _write_frequency_source(unrelated, include_complex_modulus=True)
    monkeypatch.chdir(tmp_path)
    study_model = _frequency_study_model()
    study_model["figure_queue"] = [study_model["figure_queue"][0]]

    queue = _rheology_frequency_figure_queue(
        {
            "rule_id": "rheology_frequency_sweep",
            "template": "point_line",
            "study_model": study_model,
        }
    )

    assert [item["id"] for item in queue] == ["storage_modulus_vs_frequency"]


def test_frequency_queue_fallback_does_not_read_unrelated_invalid_workbook(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    (tmp_path / "unrelated.xlsx").write_bytes(b"not an xlsx workbook")
    monkeypatch.chdir(tmp_path)
    study_model = _frequency_study_model()
    study_model["figure_queue"] = [study_model["figure_queue"][0]]

    queue = _rheology_frequency_figure_queue(
        {
            "rule_id": "rheology_frequency_sweep",
            "template": "point_line",
            "study_model": study_model,
        }
    )

    assert [item["id"] for item in queue] == ["storage_modulus_vs_frequency"]


def test_frequency_source_metric_change_invalidates_persisted_plan(
    tmp_path: Path,
) -> None:
    source = tmp_path / "frequency.xlsx"
    _write_frequency_source(source, include_complex_modulus=False)
    study_model = _frequency_study_model()
    plan = resolve_figure_plan(
        rule_id="rheology_frequency_sweep",
        template="point_line",
        study_model=study_model,
        input_path=source,
        request={},
    )
    assert plan is not None
    _write_frequency_source(source, include_complex_modulus=True)

    with pytest.raises(FigurePlanResolutionError) as exc_info:
        resolve_current_figure_plan(
            persisted=plan.to_payload(),
            rule_id="rheology_frequency_sweep",
            template="point_line",
            study_model=study_model,
            input_path=source,
            request={},
        )

    assert exc_info.value.reason_code == "stale_resolved_figure_plan"

    refreshed = resolve_preparation_figure_plan(
        persisted=plan.to_payload(),
        rule_id="rheology_frequency_sweep",
        template="point_line",
        study_model=study_model,
        input_path=source,
        request={},
    )

    assert refreshed is not None
    assert refreshed.plan_sha256 != plan.plan_sha256
    assert refreshed.source_sha256 != plan.source_sha256
    assert "complex_modulus_vs_frequency" in refreshed.selected_figure_ids


def test_studio_regeneration_failure_rolls_back_plan_and_primary_document(
    tmp_path: Path,
) -> None:
    from sciplot_core.studio import prepare_studio_document
    from sciplot_core.studio_core.figure_set_state import (
        _replace_studio_figure_set_path,
    )
    from sciplot_core.studio_core.studio_prepare import (
        prepare_studio_document as prepare_with_replacer,
    )

    source = tmp_path / "frequency.xlsx"
    _write_frequency_source(
        source,
        include_complex_modulus=False,
        include_loss_modulus=True,
        include_loss_factor_and_viscosity=True,
    )
    prepared = prepare_studio_document(
        source,
        output_root=tmp_path / "projects",
        rule_id="rheology_frequency_sweep",
        template="point_line",
    )
    project = Path(str(prepared["project_dir"]))
    request_path = project / "plot_request.json"
    document = project / "studio" / "document.vsz"
    spec = project / "studio" / "spec.json"
    registry = project / "studio" / "figure_set.json"
    prepared_registry = json.loads(registry.read_text(encoding="utf-8"))
    outcomes = prepared_registry["resolved_figure_plan"]["outcomes"]
    assert prepared_registry["version"] == 2
    assert [
        entry["resolved_figure_task"] for entry in prepared_registry["figures"]
    ] == prepared_registry["resolved_figure_plan"]["tasks"]
    assert outcomes
    assert {outcome["status"] for outcome in outcomes} == {"editable"}
    assert all(
        Path(artifact).is_file() and ".sciplot-figure-set-transaction-" not in artifact
        for outcome in outcomes
        for artifact in outcome["artifacts"]
    )
    before = {
        path: path.read_bytes() for path in (request_path, document, spec, registry)
    }
    request = json.loads(request_path.read_text(encoding="utf-8"))
    project_source = Path(str(request["input"]))
    if not project_source.is_absolute():
        project_source = request_path.parent / project_source
    if project_source.is_dir():
        project_source = next(project_source.rglob("*.xlsx"))
    _write_frequency_source(
        project_source,
        include_complex_modulus=True,
        include_loss_modulus=True,
        include_loss_factor_and_viscosity=True,
    )

    def fail_request_replace(staged: Path, target: Path) -> None:
        if target == request_path:
            raise OSError("injected request replacement failure")
        _replace_studio_figure_set_path(staged, target)

    with pytest.raises(OSError, match="injected request replacement failure"):
        prepare_with_replacer(
            project,
            regenerate_generated=True,
            figure_set_path_replacer=fail_request_replace,
        )

    assert {
        path: path.read_bytes() for path in (request_path, document, spec, registry)
    } == before
    assert not list(project.rglob(".sciplot-studio-prepare-*"))
    assert not list(project.rglob(".sciplot-figure-set-transaction-*"))


def test_preparation_cannot_refresh_a_persisted_plan_across_rule_boundaries(
    tmp_path: Path,
) -> None:
    source = tmp_path / "frequency.xlsx"
    _write_frequency_source(source, include_complex_modulus=False)
    plan = resolve_figure_plan(
        rule_id="rheology_frequency_sweep",
        template="point_line",
        study_model=_frequency_study_model(),
        input_path=source,
        request={},
    )
    assert plan is not None

    with pytest.raises(FigurePlanResolutionError) as exc_info:
        resolve_preparation_figure_plan(
            persisted=plan.to_payload(),
            rule_id="legacy_custom_rule",
            template="curve",
            study_model={},
            input_path=source,
            request={},
        )

    assert exc_info.value.reason_code == "stale_resolved_figure_plan"


def test_studio_publish_inventory_rejects_source_drift_after_prepare(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import sciplot_core.studio_core.publish_inventory as inventory_module

    project = tmp_path / "project"
    source = project / "frequency.xlsx"
    request_path = project / "plot_request.json"
    document = project / "studio" / "document.vsz"
    document.parent.mkdir(parents=True)
    document.write_bytes(b"prepared-document")
    _write_frequency_source(source, include_complex_modulus=False)
    study_model = _frequency_study_model()
    plan = resolve_figure_plan(
        rule_id="rheology_frequency_sweep",
        template="point_line",
        study_model=study_model,
        input_path=source,
        request={},
    )
    assert plan is not None
    request_path.write_text(
        json.dumps(
            {
                "rule_id": "rheology_frequency_sweep",
                "template": "point_line",
                "input": str(source),
                "study_model": study_model,
                "resolved_figure_plan": plan.to_payload(),
            }
        ),
        encoding="utf-8",
    )
    workbook = load_workbook(source)
    workbook.active.cell(row=4, column=2, value=999.0)
    workbook.save(source)
    monkeypatch.setattr(
        inventory_module,
        "_verify_exact_current_export_binding",
        lambda **_kwargs: None,
    )
    collect_calls: list[bool] = []
    monkeypatch.setattr(
        inventory_module,
        "_collect_figure_documents",
        lambda **_kwargs: collect_calls.append(True),
    )

    with pytest.raises(RuntimeError, match="stale_resolved_figure_plan"):
        inventory_module.prepare_studio_export_inventory(
            project_dir=project,
            request_path=request_path,
            document_path=document,
            exports=[],
            export_document_sha256="prepared-hash",
        )

    assert collect_calls == []
    assert not (project / "runs").exists()


@pytest.mark.parametrize("request_mutation", ["drop_plan", "drop_rule"])
def test_studio_publish_inventory_cannot_downgrade_prepared_plan(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    request_mutation: str,
) -> None:
    import sciplot_core.studio_core.publish_inventory as inventory_module

    project = tmp_path / "project"
    source = project / "frequency.xlsx"
    request_path = project / "plot_request.json"
    document = project / "studio" / "document.vsz"
    document.parent.mkdir(parents=True)
    document.write_bytes(b"prepared-document")
    _write_frequency_source(source, include_complex_modulus=False)
    study_model = _frequency_study_model()
    plan = resolve_figure_plan(
        rule_id="rheology_frequency_sweep",
        template="point_line",
        study_model=study_model,
        input_path=source,
        request={},
    )
    assert plan is not None
    request = {
        "rule_id": "rheology_frequency_sweep",
        "template": "point_line",
        "input": str(source),
        "study_model": study_model,
        "resolved_figure_plan": plan.to_payload(),
    }
    request.pop(
        "resolved_figure_plan" if request_mutation == "drop_plan" else "rule_id"
    )
    request_path.write_text(json.dumps(request), encoding="utf-8")
    monkeypatch.setattr(
        inventory_module,
        "_verify_exact_current_export_binding",
        lambda **_kwargs: None,
    )

    with pytest.raises(
        RuntimeError,
        match=(
            "prepared_resolved_figure_plan_required"
            if request_mutation == "drop_plan"
            else "stale_resolved_figure_plan"
        ),
    ):
        inventory_module.prepare_studio_export_inventory(
            project_dir=project,
            request_path=request_path,
            document_path=document,
            exports=[],
            export_document_sha256="prepared-hash",
        )


def test_studio_run_source_binding_checks_live_and_archived_bytes(
    tmp_path: Path,
) -> None:
    source = tmp_path / "frequency.xlsx"
    archive = tmp_path / "run" / "raw" / "frequency.xlsx"
    _write_frequency_source(source, include_complex_modulus=False)
    archive.parent.mkdir(parents=True)
    archive.write_bytes(source.read_bytes())
    plan = resolve_figure_plan(
        rule_id="rheology_frequency_sweep",
        template="point_line",
        study_model=_frequency_study_model(),
        input_path=source,
        request={},
    )
    assert plan is not None
    sources = StudioRunSources(
        input_path=source,
        raw_archive={"path": str(archive)},
        existing_transform_ledger=None,
        snapshot_sources=[],
        snapshot_source=None,
        processed_source=None,
        semantic={},
        metric_source=None,
        analysis_metrics=[],
    )

    verify_studio_run_source_binding(plan, sources)

    source.write_bytes(b"changed-live-source")
    with pytest.raises(RuntimeError, match="source changed"):
        verify_studio_run_source_binding(plan, sources)

    source.write_bytes(archive.read_bytes())
    archive.write_bytes(b"changed-archive")
    with pytest.raises(RuntimeError, match="source changed"):
        verify_studio_run_source_binding(plan, sources)
