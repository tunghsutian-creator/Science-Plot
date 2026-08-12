from __future__ import annotations

import json
from dataclasses import replace
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

import sciplot_core.figure_plan.source_binding as source_binding_hashes
from sciplot_core import workflow
from sciplot_core._paths import resolve_fixture_path
from sciplot_core.figure_plan import CartesianMetricBinding, FigurePlanResolutionError
from sciplot_core.materials_rules import get_rule
from sciplot_core.readiness.rule_contract import rule_contract_hashes
from sciplot_core.semantic import prepare_semantic_source
from sciplot_core.semantic_sources.rheology_temperature_domain import (
    ResolvedRheologyTemperatureDomain,
)
from sciplot_core.semantic_sources.scientific_source import (
    ResolvedScientificSource,
    ScientificSourceResolutionError,
    resolve_scientific_source,
)
from sciplot_core.studio_core.prepare_generated import generate_studio_document
from sciplot_core.studio_core.publish_sources import (
    StudioRunSources,
    verify_studio_run_source_binding,
)
from sciplot_core.study_model import (
    STUDY_MODEL_KIND,
    STUDY_MODEL_VERSION,
    experiment_recommendation_payload,
)
from sciplot_core.workflow import request_rendering
from sciplot_core.workflow.scientific_source_resolution import (
    bind_workflow_semantic_render_options,
)
from sciplot_core.workflow.request_rendering import RequestRenderResult
from sciplot_core.workflow.route_intent import resolve_workflow_route_intent
from sciplot_core.workflow.source_binding import (
    verify_workflow_figure_plan_source_binding,
)


def _write_source(path: Path) -> None:
    path.write_text(
        "Time,Shear Stress\n"
        "s,Pa\n"
        "sample A,sample A\n"
        "0.13,10\n"
        "0.29,8\n"
        "0.47,6\n",
        encoding="utf-8",
    )


def test_workflow_rule_render_contract_overrides_only_nonexplicit_defaults() -> None:
    request = {
        "render_options": {
            "size": "60x55",
            "style_preset": "nature",
            "y_label_override": "User label",
        },
        "explicit_render_option_keys": [],
    }
    semantic = {
        "render_options": {
            "size": "120x55",
            "yscale": "log",
            "line_alpha": 0.8,
            "legend_edge_padding_mm": 1.0,
        },
        "template": "curve",
        "axis_plan": {
            "x": {"display_label": "Source x"},
            "y": {"display_label": "Source y"},
        },
    }
    plan = SimpleNamespace(selection_policy="registered_single_curve")

    inherited = bind_workflow_semantic_render_options(
        request=request,
        semantic=semantic,
        figure_plan=plan,
    )
    assert inherited["render_options"] == {
        "size": "120x55",
        "style_preset": "nature",
        "yscale": "log",
        "x_label_override": "Source x",
        "y_label_override": "Source y",
    }

    explicit_request = {
        **request,
        "explicit_render_option_keys": ["size", "y_label_override"],
    }
    explicit = bind_workflow_semantic_render_options(
        request=explicit_request,
        semantic=semantic,
        figure_plan=plan,
    )
    assert explicit["render_options"]["size"] == "60x55"
    assert explicit["render_options"]["y_label_override"] == "User label"
    assert explicit["render_options"]["x_label_override"] == "Source x"


def test_workflow_single_curve_labels_come_from_the_resolved_source() -> None:
    transform = SimpleNamespace(
        contract=SimpleNamespace(
            output={
                "x_label": "Wavenumber",
                "x_unit": "cm^-1",
                "y_label": "Transmittance",
                "y_unit": "%",
            }
        )
    )
    resolved_source = SimpleNamespace(transform=transform)

    bound = bind_workflow_semantic_render_options(
        request={"render_options": {}, "explicit_render_option_keys": []},
        semantic={
            "template": "stacked_curve",
            "axis_plan": {
                "x": {"display_label": "Static x"},
                "y": {"display_label": "Spectral response"},
            },
        },
        figure_plan=SimpleNamespace(selection_policy="registered_single_curve"),
        resolved_scientific_source=resolved_source,
    )

    assert bound["render_options"]["x_label_override"] == "Wavenumber (cm⁻¹)"
    assert bound["render_options"]["y_label_override"] == "Transmittance (%)"


def _frequency_study_model() -> dict[str, object]:
    recommendation = experiment_recommendation_payload(
        rule_id="rheology_frequency_sweep"
    )
    return {
        "kind": STUDY_MODEL_KIND,
        "version": STUDY_MODEL_VERSION,
        "figure_queue": [
            {**figure, "order": order, "status": "planned"}
            for order, figure in enumerate(
                recommendation["figure_queue"],
                start=1,
            )
        ],
    }


def _write_frequency_source(path: Path) -> None:
    path.write_text(
        "Angular Frequency,Storage Modulus,Loss Modulus\n"
        "rad/s,Pa,Pa\n"
        "1,100,80\n"
        "10,90,70\n",
        encoding="utf-8",
    )


def _paired_source_binding_snapshot(
    tmp_path: Path,
    *,
    rule_id: str = "tga_curve",
) -> tuple[Path, Path, ResolvedScientificSource]:
    rule = get_rule(rule_id)
    fixture = resolve_fixture_path(str(rule.fixture_path or ""))
    live_source = tmp_path / fixture.name
    archive_source = tmp_path / "archive" / fixture.name
    archive_source.parent.mkdir()
    live_source.write_bytes(fixture.read_bytes())
    archive_source.write_bytes(fixture.read_bytes())
    resolved = resolve_scientific_source(
        live_source,
        rule_id=rule.rule_id,
        request={},
        template=rule.template,
    )
    assert resolved is not None
    assert resolved.figure_plan is not None
    return live_source, archive_source, resolved


def _spy_source_tree_hashes(
    monkeypatch: pytest.MonkeyPatch,
) -> list[Path]:
    calls: list[Path] = []
    original = source_binding_hashes.source_tree_sha256

    def counted(source: Path | None) -> str | None:
        assert source is not None
        calls.append(source.expanduser().resolve())
        return original(source)

    monkeypatch.setattr(source_binding_hashes, "source_tree_sha256", counted)
    return calls


@pytest.mark.parametrize("rule_id", ("tga_curve", "dsc_curve"))
def test_workflow_source_binding_hashes_only_archive_for_paired_snapshot(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    rule_id: str,
) -> None:
    live_source, archive_source, resolved = _paired_source_binding_snapshot(
        tmp_path,
        rule_id=rule_id,
    )
    hash_calls = _spy_source_tree_hashes(monkeypatch)

    verify_workflow_figure_plan_source_binding(
        resolved.figure_plan,
        input_path=live_source,
        raw_archive={"path": str(archive_source)},
        resolved_scientific_source=resolved,
    )

    assert hash_calls == [archive_source.resolve()]


def test_studio_dsc_source_binding_uses_the_shared_source_tree_hash(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    live_source, archive_source, resolved = _paired_source_binding_snapshot(
        tmp_path,
        rule_id="dsc_curve",
    )
    assert resolved.figure_plan is not None
    sources = StudioRunSources(
        input_path=live_source,
        raw_archive={"path": str(archive_source)},
        existing_transform_ledger=None,
        snapshot_sources=[],
        snapshot_source=None,
        processed_source=None,
        semantic={},
        metric_source=None,
        analysis_metrics=[],
    )
    hash_calls = _spy_source_tree_hashes(monkeypatch)

    verify_studio_run_source_binding(resolved.figure_plan, sources)

    assert hash_calls == [live_source.resolve(), archive_source.resolve()]


def test_workflow_source_binding_without_snapshot_hashes_live_and_archive(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    live_source, archive_source, resolved = _paired_source_binding_snapshot(tmp_path)
    hash_calls = _spy_source_tree_hashes(monkeypatch)

    verify_workflow_figure_plan_source_binding(
        resolved.figure_plan,
        input_path=live_source,
        raw_archive={"path": str(archive_source)},
    )

    assert hash_calls == [live_source.resolve(), archive_source.resolve()]


def test_workflow_source_binding_rejects_drifted_archive_for_paired_snapshot(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    live_source, archive_source, resolved = _paired_source_binding_snapshot(tmp_path)
    archive_source.write_text("archive drift\n", encoding="utf-8")
    hash_calls = _spy_source_tree_hashes(monkeypatch)

    with pytest.raises(RuntimeError, match="Workflow source changed"):
        verify_workflow_figure_plan_source_binding(
            resolved.figure_plan,
            input_path=live_source,
            raw_archive={"path": str(archive_source)},
            resolved_scientific_source=resolved,
        )

    assert hash_calls == [archive_source.resolve()]


def test_scientific_source_resolution_uses_the_rule_adapter(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import sciplot_core.semantic_sources.scientific_source as source_resolution

    rule = get_rule("rheology_stress_relaxation")
    missing_source = tmp_path / "source-does-not-need-to-exist.csv"
    assert (
        source_resolution.resolve_scientific_source(
            missing_source,
            rule_id=None,
            request={},
            template=rule.template,
        )
        is None
    )
    assert (
        source_resolution.resolve_scientific_source(
            missing_source,
            rule_id="not_a_registered_rule",
            request={},
            template=rule.template,
        )
        is None
    )
    monkeypatch.setattr(
        source_resolution,
        "get_rule",
        lambda _rule_id: replace(rule, fixture_status="pending"),
    )
    assert (
        source_resolution.resolve_scientific_source(
            missing_source,
            rule_id=rule.rule_id,
            request={},
            template=rule.template,
        )
        is None
    )
    monkeypatch.setattr(
        source_resolution,
        "get_rule",
        lambda _rule_id: replace(
            rule,
            scientific_source_adapter=None,
            figure_plan_adapter=None,
        ),
    )
    assert (
        source_resolution.resolve_scientific_source(
            missing_source,
            rule_id=rule.rule_id,
            request={},
            template=rule.template,
        )
        is None
    )


def test_scientific_source_adapter_is_internal_execution_metadata() -> None:
    original = get_rule("rheology_temperature_sweep")
    rerouted = replace(original, scientific_source_adapter=None)

    assert rerouted.to_payload() == original.to_payload()
    assert rule_contract_hashes(rerouted) == rule_contract_hashes(original)


def test_stress_source_binds_one_transform_to_one_source_hashed_plan(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = tmp_path / "relaxation.csv"
    _write_source(source)
    rule = get_rule("rheology_stress_relaxation")
    import sciplot_core.semantic_sources.stress_relaxation_transform as transform

    original = transform.resolve_stress_relaxation_transform
    calls: list[Path] = []
    snapshots: list[object] = []

    def counted(source_path: Path, *, series_order: object = None):
        calls.append(source_path.expanduser().resolve())
        snapshot = original(source_path, series_order=series_order)
        snapshots.append(snapshot)
        return snapshot

    monkeypatch.setattr(transform, "resolve_stress_relaxation_transform", counted)

    resolved = resolve_scientific_source(
        source,
        rule_id=rule.rule_id,
        request={},
        template=rule.template,
    )

    assert resolved is not None
    assert resolved.transform is snapshots[0]
    assert calls == [source.resolve()]
    assert resolved.source_sha256 is not None
    plan = resolved.figure_plan
    assert plan is not None
    assert plan.source_sha256 == resolved.source_sha256
    assert plan.selection_policy == "registered_single_curve"
    assert len(plan.tasks) == 1
    task = plan.tasks[0]
    output = resolved.transform.contract.output
    binding = CartesianMetricBinding(
        x_metric=str(output["x_metric"]),
        y_metric=str(output["y_metric"]),
    )
    sample_order = tuple(series.sample for series in resolved.transform.series)
    expected_stem = "rheology_stress_relaxation_normalized_stress_vs_time"
    assert binding == CartesianMetricBinding(
        x_metric="time",
        y_metric="normalized_stress",
    )
    assert task.metric_binding == binding
    assert task.sample_order == sample_order
    assert task.artifact_stem == expected_stem
    assert task.document_stem == expected_stem
    assert plan.primary_figure_id == expected_stem


def test_scientific_source_adapters_keep_stable_error_reasons(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import sciplot_core.figure_plan.dma_temperature_resolution as dma_resolution
    import sciplot_core.semantic_sources.stress_relaxation_transform as stress

    stress_source = tmp_path / "stress.csv"
    stress_source.write_text("source", encoding="utf-8")
    monkeypatch.setattr(
        stress,
        "resolve_stress_relaxation_transform",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(ValueError("anchor missing")),
    )
    with pytest.raises(ScientificSourceResolutionError) as stress_error:
        resolve_scientific_source(
            stress_source,
            rule_id="rheology_stress_relaxation",
            request={},
            template="curve",
        )
    assert isinstance(stress_error.value, ValueError)
    assert stress_error.value.reason_code == "stress_relaxation_transform_invalid"
    assert str(stress_error.value) == "anchor missing"

    monkeypatch.setattr(
        dma_resolution,
        "resolve_dma_temperature_source",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(
            FigurePlanResolutionError(
                "dma_temperature_source_contract_invalid",
                "bad DMA source",
            )
        ),
    )
    with pytest.raises(ScientificSourceResolutionError) as dma_error:
        resolve_scientific_source(
            tmp_path / "dma.csv",
            rule_id="dma_temperature_sweep",
            request={},
            template="point_line",
        )
    assert isinstance(dma_error.value, ValueError)
    assert dma_error.value.reason_code == "dma_temperature_source_contract_invalid"
    assert str(dma_error.value) == "bad DMA source"


def test_studio_and_workflow_each_resolve_one_scientific_source_snapshot(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = tmp_path / "relaxation.csv"
    _write_source(source)
    import sciplot_core.semantic_sources.prepare_rheology as preparation
    import sciplot_core.semantic_sources.stress_relaxation_transform as transform

    original = transform.resolve_stress_relaxation_transform
    calls: list[Path] = []
    contracts: list[dict[str, object]] = []

    def counted(source_path: Path, *, series_order: object = None):
        calls.append(source_path.expanduser().resolve())
        snapshot = original(source_path, series_order=series_order)
        contracts.append(snapshot.contract.to_payload())
        return snapshot

    monkeypatch.setattr(transform, "resolve_stress_relaxation_transform", counted)
    monkeypatch.setattr(
        preparation,
        "resolve_stress_relaxation_transform",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(
            AssertionError("semantic preparation reparsed the source")
        ),
    )

    project_dir = tmp_path / "studio_project"
    project_dir.mkdir()
    studio_request = project_dir / "plot_request.json"
    studio_request.write_text(
        json.dumps(
            {
                "input": str(source),
                "rule_id": "rheology_stress_relaxation",
                "template": "curve",
                "explicit_template_selection": True,
                "explicit_render_option_keys": [],
            }
        ),
        encoding="utf-8",
    )

    prepared = generate_studio_document(
        project_dir=project_dir,
        request_path=studio_request,
        rule_id=None,
        template=None,
        project_name=None,
    )

    assert Path(prepared["document"]).is_file()
    assert calls == [source.resolve()]
    persisted_request = json.loads(studio_request.read_text(encoding="utf-8"))
    studio_contract = next(
        step["parameters"]["scientific_transform"]
        for step in persisted_request["transform_ledger"]["steps"]
        if "scientific_transform" in step.get("parameters", {})
    )
    assert studio_contract == contracts[0]

    workflow_request = tmp_path / "workflow_request.json"
    workflow_request.write_text(
        json.dumps(
            {
                "recipe": "auto",
                "input": str(source),
                "output": str(tmp_path / "workflow_output"),
                "rule_id": "rheology_stress_relaxation",
                "exports": ["pdf", "tiff_300"],
            }
        ),
        encoding="utf-8",
    )

    manifest = workflow.run_request(workflow_request)

    assert calls == [source.resolve(), source.resolve()]
    runtime_contract = next(
        step["parameters"]["scientific_transform"]
        for step in manifest["transform_ledger"]["steps"]
        if "scientific_transform" in step.get("parameters", {})
    )
    assert runtime_contract == studio_contract
    assert runtime_contract == contracts[1]
    assert manifest["result"]["veusz_documents"]
    assert manifest["result"]["outputs"]


def test_rheology_temperature_snapshot_drives_plan_and_prepare_without_reparse(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    rule = get_rule("rheology_temperature_sweep")
    source = resolve_fixture_path(str(rule.fixture_path or ""))
    import sciplot_core.semantic as semantic_module
    import sciplot_core.semantic_sources.prepare_rheology as preparation
    import sciplot_core.semantic_sources.rheology_temperature_domain as domain_module

    original_reader = domain_module._read_rheology_temperature_comparison_samples
    reader_calls: list[Path] = []

    def counted_reader(source_path: Path):
        reader_calls.append(source_path.expanduser().resolve())
        return original_reader(source_path)

    monkeypatch.setattr(
        domain_module,
        "_read_rheology_temperature_comparison_samples",
        counted_reader,
    )
    resolved = resolve_scientific_source(
        source,
        rule_id=rule.rule_id,
        request={"replicate_mode": "mean"},
        template=rule.template,
    )
    assert resolved is not None
    assert resolved.figure_plan is not None
    assert resolved.transform is None
    domain = resolved.require_domain(ResolvedRheologyTemperatureDomain)
    assert reader_calls == [source.resolve()]

    def no_reparse(*_args: object, **_kwargs: object) -> None:
        raise AssertionError("semantic preparation reparsed the temperature source")

    monkeypatch.setattr(
        preparation,
        "_read_rheology_temperature_comparison_samples",
        no_reparse,
    )
    monkeypatch.setattr(
        preparation,
        "_read_confirmed_rheology_sweep_samples",
        no_reparse,
    )
    monkeypatch.setattr(
        preparation,
        "_coalesce_replicate_sweep_samples",
        no_reparse,
    )
    monkeypatch.setattr(preparation, "_ordered_sweep_samples", no_reparse)
    monkeypatch.setattr(
        semantic_module,
        "source_tree_sha256",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(
            AssertionError("semantic preparation rehashed the resolved source")
        ),
    )

    prepared = prepare_semantic_source(
        source,
        output_dir=tmp_path / "prepared_temperature",
        semantic={"semantic_family": rule.rule_id, "rule_id": rule.rule_id},
        replicate_mode="mean",
        resolved_scientific_source=resolved,
    )

    expected_order = domain.facts.sample_order
    assert reader_calls == [source.resolve()]
    assert all(
        task.sample_order == expected_order for task in resolved.figure_plan.tasks
    )
    parameters = prepared["transform_steps"][0]["parameters"]
    assert tuple(parameters["output_sample_labels"]) == expected_order
    attestation = prepared["source_attestation"]
    assert attestation.source_tree_sha256_before == resolved.source_sha256
    assert attestation.source_tree_sha256_after == resolved.source_sha256


def test_rheology_frequency_snapshot_drives_plan_and_prepare_without_reparse(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = tmp_path / "frequency"
    source.mkdir()
    raw = source / "sample_a.csv"
    _write_frequency_source(raw)
    (source / "derived.xlsx").write_bytes(b"not a parser-selected source")
    rule = get_rule("rheology_frequency_sweep")
    import sciplot_core.semantic_sources.prepare_rheology as preparation
    import sciplot_core.semantic_sources.rheology_sweep_domain as domain_module

    original_reader = domain_module._read_rheology_frequency_comparison_samples
    reader_calls: list[Path] = []

    def counted_reader(source_path: Path):
        reader_calls.append(source_path.expanduser().resolve())
        return original_reader(source_path)

    monkeypatch.setattr(
        domain_module,
        "_read_rheology_frequency_comparison_samples",
        counted_reader,
    )
    resolved = resolve_scientific_source(
        source,
        rule_id=rule.rule_id,
        request={},
        template=rule.template,
        study_model=_frequency_study_model(),
    )

    assert resolved is not None
    assert resolved.figure_plan is not None
    domain = resolved.require_domain(domain_module.ResolvedRheologySweepDomain)
    assert reader_calls == [source.resolve()]
    assert domain.selected_sources == (raw.resolve(),)
    assert all(
        task.sample_order == domain.facts.sample_order
        for task in resolved.figure_plan.tasks
    )

    def no_reparse(*_args: object, **_kwargs: object) -> None:
        raise AssertionError("semantic preparation reparsed the frequency source")

    monkeypatch.setattr(
        preparation,
        "_read_rheology_frequency_comparison_samples",
        no_reparse,
    )
    monkeypatch.setattr(
        preparation,
        "_read_confirmed_rheology_sweep_samples",
        no_reparse,
    )
    monkeypatch.setattr(
        preparation,
        "_coalesce_replicate_sweep_samples",
        no_reparse,
    )
    monkeypatch.setattr(preparation, "_ordered_sweep_samples", no_reparse)

    prepared = prepare_semantic_source(
        source,
        output_dir=tmp_path / "prepared_frequency",
        semantic={
            "semantic_family": rule.semantic_family,
            "rule_id": rule.rule_id,
        },
        resolved_scientific_source=resolved,
    )

    assert reader_calls == [source.resolve()]
    parameters = prepared["transform_steps"][0]["parameters"]
    assert tuple(parameters["output_sample_labels"]) == domain.facts.sample_order
    processed = Path(str(prepared["processed_source"]))
    workbook = load_workbook(processed, read_only=True)
    headers = tuple(cell.value for cell in next(workbook.active.iter_rows()))
    workbook.close()
    assert "Complex Modulus" in headers


def test_scientific_direct_route_uses_semantic_materialization(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    render_request = {
        "template": "curve",
        "rule_id": "rheology_stress_relaxation",
    }
    source = tmp_path / "relaxation.csv"
    _write_source(source)
    resolved = resolve_scientific_source(
        source,
        rule_id="rheology_stress_relaxation",
        request=render_request,
        template="curve",
    )
    assert resolved is not None
    route = resolve_workflow_route_intent(render_request)
    sentinel = RequestRenderResult(
        route_intent=route,
        final_recipe=render_request.get("recipe"),
        result={"kind": "semantic_sentinel"},
        plotted_data_source=source,
    )
    received: list[tuple[object, object]] = []

    def fake_semantic(**kwargs: object) -> RequestRenderResult:
        received.append(
            (
                kwargs["resolved_scientific_source"],
                kwargs["selected_figure_plan"],
            )
        )
        return sentinel

    monkeypatch.setattr(
        request_rendering,
        "_render_semantic_plan_request",
        fake_semantic,
    )
    monkeypatch.setattr(
        request_rendering,
        "run_recipe",
        lambda *_args, **_kwargs: pytest.fail(
            "scientific request bypassed semantic materialization"
        ),
    )

    rendered = request_rendering.execute_request_render(
        request=render_request,
        route_intent=route,
        semantic={
            "rule_id": "rheology_stress_relaxation",
            "semantic_family": "rheology_stress_relaxation",
            "template": "curve",
        },
        study_model={},
        input_path=source,
        output_dir=tmp_path / "output",
        base_dir=tmp_path,
        transform_steps=[],
        resolved_scientific_source=resolved,
        _resolved_figure_plan=resolved.figure_plan,
    )

    assert rendered is sentinel
    assert len(received) == 1
    assert received[0][0] is resolved
    assert received[0][1] is resolved.figure_plan


def test_scientific_named_recipe_with_plan_fails_before_materialization(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = tmp_path / "relaxation.csv"
    _write_source(source)
    request = {
        "recipe": "stress_relaxation",
        "rule_id": "rheology_stress_relaxation",
    }
    resolved = resolve_scientific_source(
        source,
        rule_id="rheology_stress_relaxation",
        request=request,
        template="curve",
    )
    assert resolved is not None
    assert resolved.figure_plan is not None
    route = resolve_workflow_route_intent(request)
    unexpected_calls: list[str] = []

    def unexpected(route_name: str) -> None:
        unexpected_calls.append(route_name)
        pytest.fail(f"unsupported named FigurePlan reached {route_name}")

    monkeypatch.setattr(
        request_rendering,
        "_render_semantic_plan_request",
        lambda **_kwargs: unexpected("semantic preparation"),
    )
    monkeypatch.setattr(
        request_rendering,
        "_render_legacy_recipe_request",
        lambda **_kwargs: unexpected("legacy recipe execution"),
    )
    transform_steps: list[dict[str, object]] = []
    output_dir = tmp_path / "must_not_exist"

    with pytest.raises(
        ValueError,
        match="workflow_recipe_figure_plan_unsupported",
    ):
        request_rendering.execute_request_render(
            request=request,
            route_intent=route,
            semantic={
                "rule_id": "rheology_stress_relaxation",
                "semantic_family": "rheology_stress_relaxation",
                "template": "curve",
            },
            study_model={},
            input_path=source,
            output_dir=output_dir,
            base_dir=tmp_path,
            transform_steps=transform_steps,
            resolved_scientific_source=resolved,
            _resolved_figure_plan=resolved.figure_plan,
        )

    assert unexpected_calls == []
    assert transform_steps == []
    assert not output_dir.exists()
